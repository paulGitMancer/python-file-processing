# -*- coding: utf-8 -*-
"""
Created on Mon Sep 15 14:37:31 2025

@author: Paul
"""
#############################################################################
## Simple script I created to populate an Excel workbook with sample data. ##
#############################################################################

#Import Library
import openpyxl

#Load Workbook and Active Sheet
wb_obj = openpyxl.load_workbook ('C:/Users/Paul/Downloads/python-input.xlsx')
sheet_obj = wb_obj.active

#Identify Header Row
header_row = 4

#Identify Max Number of Rows/Columns
row = sheet_obj.max_row
column = sheet_obj.max_column

#Identify Current Working Row
myrow = 5

myrange = range (5,row+1)
for i in myrange:
    
    #Inputs

    #Employee Name
    cell_obj = sheet_obj.cell(row=i,column=1)
    emp_name = cell_obj.value
    print(emp_name)
    
    #Department
    cell_obj = sheet_obj.cell(row=i,column=2)
    dept = cell_obj.value

    #HI Enrollment
    cell_obj = sheet_obj.cell(row=i,column=3)
    HI_enrollment = cell_obj.value

    #Pay Rate
    cell_obj = sheet_obj.cell(row=i,column=4)
    pay_rate = cell_obj.value

    #Pay Basis
    cell_obj = sheet_obj.cell(row=i,column=5)
    pay_basis = cell_obj.value

    #Hours Worked
    cell_obj = sheet_obj.cell(row=i,column=6)
    hours_worked = cell_obj.value

    #Logic Statements
    if hours_worked > 40:
        reg_hours = 40
        OT_hours = hours_worked - 40 
    else:
        reg_hours = hours_worked
        OT_hours = 0
        
    if pay_basis == "Salary":
        reg_pay = pay_rate
        OT_pay = 0
    else:
        reg_pay = pay_rate*reg_hours
        OT_pay = pay_rate*OT_hours*1.5
    
    gross_pay = reg_pay+OT_pay
    
    fed_tax = gross_pay*0.3
    
    state_tax = gross_pay*0.04
    
    local_tax = gross_pay*0.03
    
    FICA = gross_pay*0.0625
    
    if HI_enrollment == "F":
        HI_deduction = 250
    elif HI_enrollment == "I":
        HI_deduction = 100
    else: 
        HI_deduction = 25
    
    total_deductions = fed_tax+state_tax+local_tax+FICA+HI_deduction
    
    net_pay = gross_pay-total_deductions

    #Outputs
    
    cell_obj = sheet_obj.cell(row=i,column=7)
    cell_obj.value = reg_hours
    
    cell_obj = sheet_obj.cell(row=i,column=8)
    cell_obj.value = OT_hours
    
    cell_obj = sheet_obj.cell(row=i,column=9)
    cell_obj.value = reg_pay
    
    cell_obj = sheet_obj.cell(row=i,column=10)
    cell_obj.value = OT_pay
    
    cell_obj = sheet_obj.cell(row=i,column=11)
    cell_obj.value = gross_pay
    
    cell_obj = sheet_obj.cell(row=i,column=12)
    cell_obj.value = fed_tax
    
    cell_obj = sheet_obj.cell(row=i,column=13)
    cell_obj.value = state_tax
    
    cell_obj = sheet_obj.cell(row=i,column=14)
    cell_obj.value = local_tax
    
    cell_obj = sheet_obj.cell(row=i,column=15)
    cell_obj.value = FICA
    
    cell_obj = sheet_obj.cell(row=i,column=16)
    cell_obj.value = HI_deduction
    
    cell_obj = sheet_obj.cell(row=i,column=17)
    cell_obj.value = total_deductions
    
    cell_obj = sheet_obj.cell(row=i,column=18)
    cell_obj.value = net_pay
    
wb_obj.save('C:/Users/Paul/Downloads/python-output.xlsx')